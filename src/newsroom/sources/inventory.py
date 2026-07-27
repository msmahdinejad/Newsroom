"""Optional extended XLSX inventory importer and source registry.

Parses an operator-supplied workbook (``All Sources`` sheet is authoritative),
preserves every workbook row in the ``source_inventory`` table with a stable
normalized identity (independent of display name), validates each row, and
activates accessible sources into the existing ``sources`` collector registry.

Design guarantees:
  * all workbook rows are accounted for (active / inactive / invalid);
  * no row disappears silently;
  * repeated import is idempotent by ``stable_identity`` (no duplicates);
  * source identity does not depend on display name;
  * disabling a source never removes its historical raw items;
  * invalid rows are reported individually and do not stop the import;
  * existing source-state names are reused (configured/healthy/degraded/
    unavailable on ``sources``; active/inactive/invalid on the inventory).

No credentials, cookies, or session material are stored or logged.
"""

from __future__ import annotations

import hashlib
import os
import shutil
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import urlparse, urlunparse

from sqlalchemy.orm import Session

from newsroom.logging import get_logger
from newsroom.storage.models import Source, SourceInventory

logger = get_logger(__name__)

# Extended inventory schema.
AUTHORITATIVE_SHEET = "All Sources"

# Workbook column header names (exact strings from the workbook).
COL_ID = "ID"
COL_PLATFORM = "Platform"
COL_TYPE = "Type"
COL_NAME = "Name"
COL_HANDLE = "Handle / ID"
COL_URL = "Direct URL"
COL_TOPIC = "Primary Topic"
COL_TAGS = "Tags"
COL_LANG = "Language"
COL_CONTENT_MODE = "Content Mode"
COL_SPEED = "Speed 1-5"
COL_INFORMAL = "Informal 1-5"
COL_NOISE = "Noise 1-5"
COL_COMMUNITY = "Community?"
COL_OPENSOURCE_API = "Open-source/API?"
COL_RISK = "Risk"
COL_VERIFICATION = "Verification"
COL_DISCOVERY = "Discovery Source"
COL_TIER = "Tier"
COL_COVERAGE = "Coverage Score"

# Workbook location (relative to repo root or absolute).
SOURCE_WORKBOOK_ENV = "NEWSROOM_SOURCE_WORKBOOK"
IMPORT_DEST = Path("config/import/source-radar.xlsx")


# ── Stable identity ───────────────────────────────────────────────


def _norm_url(url: str) -> str:
    """Normalize a URL for stable identity: lowercase host, strip tracking
    params/fragment, strip trailing slash, default https scheme."""
    if not url:
        return ""
    u = url.strip()
    if not u:
        return ""
    if not u.startswith(("http://", "https://")):
        # t.me, etc. — leave as-is for platform-specific normalizers
        return u.lower().rstrip("/")
    try:
        parsed = urlparse(u)
    except Exception:
        return u.lower().rstrip("/")
    host = (parsed.hostname or "").lower()
    path = parsed.path or ""
    # Normalize root/empty path and strip trailing slash for identity stability.
    if path in ("", "/"):
        path = ""
    elif path.endswith("/"):
        path = path.rstrip("/")
    # Drop tracking query params; keep nothing for identity stability.
    rebuilt = urlunparse(("https", host, path, "", "", ""))
    return rebuilt


def _tg_handle(handle: str | None, url: str) -> str:
    h = (handle or "").strip().lstrip("@").lower()
    if not h and url:
        p = urlparse(url if url.startswith("http") else f"https://{url}")
        parts = [x for x in p.path.split("/") if x]
        # t.me/<handle> or t.me/s/<handle>
        if parts and parts[0] == "s" and len(parts) > 1:
            h = parts[1]
        elif parts:
            h = parts[0]
    return h


def _reddit_sub(handle: str | None, url: str) -> str:
    h = (handle or "").strip().lower()
    if h.startswith("r/"):
        h = h[2:]
    if not h and url:
        p = urlparse(url)
        parts = [x for x in p.path.split("/") if x]
        if parts and parts[0] == "r" and len(parts) > 1:
            h = parts[1]
    return h


def _x_handle(handle: str | None, url: str) -> str:
    h = (handle or "").strip().lstrip("@").lower()
    if not h and url:
        p = urlparse(url)
        parts = [x for x in p.path.split("/") if x]
        if parts:
            h = parts[0]
    return h


def _github_repo(url: str) -> str:
    """Return 'owner/repo' lowercased for a repo URL, or '' for non-repo."""
    if not url:
        return ""
    p = urlparse(url)
    host = (p.hostname or "").lower()
    if host not in {"github.com", "www.github.com"}:
        return ""
    parts = [x for x in p.path.split("/") if x]
    if len(parts) >= 2:
        return f"{parts[0].lower()}/{parts[1].lower()}"
    return ""


def _yt_handle(handle: str | None, url: str) -> str:
    h = (handle or "").strip().lstrip("@").lower()
    if not h and url:
        p = urlparse(url)
        host = (p.hostname or "").lower()
        parts = [x for x in p.path.split("/") if x]
        if host in {"www.youtube.com", "youtube.com"} and parts:
            if parts[0].startswith("@"):
                h = parts[0].lstrip("@").lower()
            elif parts[0] == "channel" and len(parts) > 1:
                h = f"channel:{parts[1].lower()}"
            elif parts[0] == "c" and len(parts) > 1:
                h = parts[1].lower()
    return h


def stable_identity_for(platform: str, handle: str | None, url: str) -> str:
    """Deterministic stable identity independent of display name."""
    if platform == "Telegram":
        return hashlib.sha256(f"telegram:{_tg_handle(handle, url)}".encode()).hexdigest()
    if platform == "Reddit":
        return hashlib.sha256(f"reddit:{_reddit_sub(handle, url)}".encode()).hexdigest()
    if platform == "X / Twitter":
        return hashlib.sha256(f"x:{_x_handle(handle, url)}".encode()).hexdigest()
    if platform == "GitHub":
        repo = _github_repo(url)
        return hashlib.sha256(f"github:{repo or _norm_url(url)}".encode()).hexdigest()
    if platform == "YouTube / Social":
        return hashlib.sha256(f"youtube:{_yt_handle(handle, url)}".encode()).hexdigest()
    # Website / Newsletter, Community, Community / Forum
    return hashlib.sha256(f"web:{_norm_url(url)}".encode()).hexdigest()


# ── Platform → mapped source type + accessibility ────────────────

# Map platforms requiring membership/owner-side auth that we do not attempt.
ACCESS_REQUIRED_TYPES: frozenset[str] = frozenset(
    {"Discord", "Slack", "Bot"}
)
PERMANENT_RUNTIME_INACTIVE_REASONS: frozenset[str] = frozenset(
    {"channel_private", "channel_unresolvable", "duplicate_identity", "handle_missing"}
)


def mapped_type_for(platform: str, workbook_type: str, url: str) -> tuple[str, str | None]:
    """Return (mapped_newsroom_type, inactive_reason_or_None).

    inactive_reason is None when the source is a candidate for activation
    (subject to validation); a non-None value means the row is inactive by
    design (access-dependent / not a repo / unsupported).
    """
    if platform == "Telegram":
        return "telegram", None
    if platform == "Reddit":
        return "reddit_subreddit", None
    if platform == "X / Twitter":
        return "x_timeline", None  # activation checks X auth availability
    if platform == "GitHub":
        if _github_repo(url):
            return "github_releases", None
        return "github_releases", "not_a_repo"  # e.g. github.com/trending
    if platform == "YouTube / Social":
        return "youtube_rss", None
    if platform in ("Community", "Community / Forum"):
        if workbook_type in ACCESS_REQUIRED_TYPES:
            return "web_page", "access_required"
        return "web_page", None
    # Website / Newsletter
    return "web_page", None


def _bool(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in {"yes", "true", "1", "y"}


def _int(val: Any) -> int:
    try:
        return int(val) if val is not None else 0
    except (ValueError, TypeError):
        return 0


def _int_opt(val: Any) -> int | None:
    try:
        return int(val) if val not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _str(val: Any) -> str:
    return str(val).strip() if val is not None else ""


def _trunc(val: str, n: int) -> str:
    return val[:n] if len(val) > n else val


def _review_level(verification: str) -> str:
    """Derive a short review-level category from the workbook Verification text.

    Fits the varchar(50) column and gives a coarse, comparable review level.
    """
    v = (verification or "").lower()
    if "official page checked" in v:
        return "official_checked"
    if "checked 2026" in v:
        return "directory_checked"
    if "known/stable" in v or "known public" in v or "known or likely public" in v:
        return "known_public"
    if "public" in v:
        return "public_unverified"
    if "found in public telegram directory" in v:
        return "directory_listed"
    return "unverified"


# ── Workbook location ─────────────────────────────────────────────


def find_workbook(
    repo_root: str | Path = ".",
    workbook_path: str | Path | None = None,
) -> Path | None:
    """Locate an explicit, environment-configured, or repo-local workbook."""
    root = Path(repo_root).resolve()
    configured = workbook_path or os.environ.get(SOURCE_WORKBOOK_ENV, "").strip()
    if configured:
        candidate = Path(configured).expanduser()
        if not candidate.is_absolute():
            candidate = root / candidate
        candidate = candidate.resolve()
        return candidate if candidate.is_file() else None

    canonical = (root / IMPORT_DEST).resolve()
    if canonical.is_file():
        return canonical

    return None


def copy_workbook_to_import_dir(src: Path, repo_root: str | Path = ".") -> Path:
    """Copy the workbook to config/import/source-radar.xlsx without changing
    the original. Returns the destination path."""
    dest_dir = Path(repo_root) / "config/import"
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / "source-radar.xlsx"
    if src.resolve() == dest.resolve():
        return dest
    shutil.copy2(src, dest)
    return dest


# ── Parsing ───────────────────────────────────────────────────────


@dataclass
class ParsedRow:
    workbook_id: int
    platform: str
    workbook_type: str
    name: str
    handle: str
    public_url: str
    topic: str
    tags: str
    language: str
    content_mode: str
    speed: int | None
    informal: int | None
    noise: int | None
    is_community: bool
    is_opensource_api: bool
    risk: str
    verification: str
    discovery_source: str
    tier: str
    coverage_score: int


def _load_workbook_rows(path: Path) -> list[list[Any]]:
    """Load the All Sources sheet rows (header + data) from the workbook."""
    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(str(path), data_only=True)
    if AUTHORITATIVE_SHEET not in wb.sheetnames:
        raise ValueError(f"workbook has no '{AUTHORITATIVE_SHEET}' sheet; sheets: {wb.sheetnames}")
    ws = wb[AUTHORITATIVE_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    return rows


def _parse_rows(rows: list[list[Any]]) -> list[ParsedRow]:
    header = [str(c) if c is not None else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    def col(row, name, default=""):
        i = idx.get(name)
        if i is None or i >= len(row):
            return default
        return row[i]

    parsed: list[ParsedRow] = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        wb_id = _int_opt(col(row, COL_ID))
        if wb_id is None:
            continue  # not a data row
        parsed.append(
            ParsedRow(
                workbook_id=wb_id,
                platform=_str(col(row, COL_PLATFORM)),
                workbook_type=_str(col(row, COL_TYPE)),
                name=_str(col(row, COL_NAME)),
                handle=_str(col(row, COL_HANDLE)),
                public_url=_str(col(row, COL_URL)),
                topic=_str(col(row, COL_TOPIC)),
                tags=_str(col(row, COL_TAGS)),
                language=_str(col(row, COL_LANG)),
                content_mode=_str(col(row, COL_CONTENT_MODE)),
                speed=_int_opt(col(row, COL_SPEED)),
                informal=_int_opt(col(row, COL_INFORMAL)),
                noise=_int_opt(col(row, COL_NOISE)),
                is_community=_bool(col(row, COL_COMMUNITY)),
                is_opensource_api=_bool(col(row, COL_OPENSOURCE_API)),
                risk=_str(col(row, COL_RISK)),
                verification=_str(col(row, COL_VERIFICATION)),
                discovery_source=_str(col(row, COL_DISCOVERY)),
                tier=_str(col(row, COL_TIER)),
                coverage_score=_int(col(row, COL_COVERAGE)),
            )
        )
    return parsed


def _validate_row(row: ParsedRow) -> tuple[str, str]:
    """Return (validation_result, validation_detail)."""
    if not row.public_url:
        return "missing_url", "Direct URL is empty"
    try:
        urlparse(row.public_url if row.public_url.startswith("http") else f"https://{row.public_url}")
    except Exception as e:
        return "invalid_url", f"unparseable URL: {e}"
    if row.platform == "Telegram":
        if not _tg_handle(row.handle, row.public_url):
            return "invalid_handle", "no resolvable Telegram handle"
    elif row.platform == "Reddit":
        if not _reddit_sub(row.handle, row.public_url):
            return "invalid_handle", "no resolvable subreddit"
    elif row.platform == "X / Twitter":
        if not _x_handle(row.handle, row.public_url):
            return "invalid_handle", "no resolvable X handle"
    elif row.platform == "YouTube / Social":
        if not _yt_handle(row.handle, row.public_url):
            return "invalid_handle", "no resolvable YouTube handle"
    elif row.platform == "GitHub":
        # Non-repo (trending) is valid as a row but inactive by design — not invalid.
        pass
    return "ok", ""


@dataclass
class ImportReport:
    workbook_path: str = ""
    total_rows: int = 0
    upserted: int = 0
    invalid: list[dict[str, Any]] = field(default_factory=list)
    duplicates: list[dict[str, Any]] = field(default_factory=list)
    platform_counts: dict[str, int] = field(default_factory=dict)
    validation_counts: dict[str, int] = field(default_factory=dict)
    duplicate_by_identity: int = 0

    def to_dict(self) -> dict[str, Any]:
        return {
            "workbook_path": self.workbook_path,
            "total_rows": self.total_rows,
            "upserted": self.upserted,
            "invalid": self.invalid,
            "duplicates": self.duplicates,
            "platform_counts": self.platform_counts,
            "validation_counts": self.validation_counts,
            "duplicate_by_identity": self.duplicate_by_identity,
        }


def import_workbook(session: Session, path: Path) -> ImportReport:
    """Idempotently import all workbook rows into source_inventory.

    Repeated import updates workbook metadata but preserves the activation
    link (source_id) and operational_state managed by ``activate_inventory``.
    """
    rows = _load_workbook_rows(path)
    parsed = _parse_rows(rows)
    report = ImportReport(workbook_path=path.name, total_rows=len(parsed))

    seen_identities: set[str] = set()
    platform_counts: dict[str, int] = {}
    validation_counts: dict[str, int] = {}

    for row in parsed:
        platform_counts[row.platform] = platform_counts.get(row.platform, 0) + 1
        sid = stable_identity_for(row.platform, row.handle, row.public_url)
        vresult, vdetail = _validate_row(row)

        is_duplicate = sid in seen_identities
        if is_duplicate:
            # Same source appears on a second workbook row. Retain the row
            # (no silent disappearance) but mark it as a duplicate so it is
            # never activated. The first occurrence is canonical.
            vresult = "duplicate"
            vdetail = "duplicate stable identity — first occurrence retained"
            report.duplicates.append(
                {"workbook_id": row.workbook_id, "name": row.name, "platform": row.platform, "identity": sid[:12]}
            )
            report.duplicate_by_identity += 1
        else:
            seen_identities.add(sid)
        validation_counts[vresult] = validation_counts.get(vresult, 0) + 1

        mapped_type, _inactive = mapped_type_for(row.platform, row.workbook_type, row.public_url)

        # Idempotency: one inventory row per workbook row (key = workbook_id).
        existing = session.query(SourceInventory).filter_by(workbook_id=row.workbook_id).first()
        if existing:
            # Update workbook metadata; preserve activation-managed fields.
            existing.platform = row.platform
            existing.workbook_type = _trunc(row.workbook_type, 50)
            existing.name = _trunc(row.name, 500)
            existing.handle = _trunc(row.handle, 255) or None
            existing.public_url = row.public_url
            existing.topic = _trunc(row.topic, 200) or None
            existing.tags = row.tags or None
            existing.language = _trunc(row.language, 30) or None
            existing.content_mode = _trunc(row.content_mode, 30) or None
            existing.review_level = _review_level(row.verification)
            existing.verification = row.verification or None
            existing.discovery_source = row.discovery_source or None
            existing.tier = _trunc(row.tier, 30) or None
            existing.coverage_score = row.coverage_score
            existing.risk = _trunc(row.risk, 30) or None
            existing.speed = row.speed
            existing.informal = row.informal
            existing.noise = row.noise
            existing.is_community = row.is_community
            existing.is_opensource_api = row.is_opensource_api
            existing.stable_identity = sid
            existing.mapped_type = mapped_type
            existing.validation_result = vresult
            existing.validation_detail = vdetail or None
            if is_duplicate:
                existing.operational_state = "duplicate"
                existing.inactive_reason = "duplicate_identity"
            report.upserted += 1
        else:
            inv = SourceInventory(
                workbook_id=row.workbook_id,
                platform=row.platform,
                workbook_type=_trunc(row.workbook_type, 50),
                name=_trunc(row.name, 500),
                handle=_trunc(row.handle, 255) or None,
                public_url=row.public_url,
                topic=_trunc(row.topic, 200) or None,
                tags=row.tags or None,
                language=_trunc(row.language, 30) or None,
                content_mode=_trunc(row.content_mode, 30) or None,
                review_level=_review_level(row.verification),
                verification=row.verification or None,
                discovery_source=row.discovery_source or None,
                tier=_trunc(row.tier, 30) or None,
                coverage_score=row.coverage_score,
                risk=_trunc(row.risk, 30) or None,
                speed=row.speed,
                informal=row.informal,
                noise=row.noise,
                is_community=row.is_community,
                is_opensource_api=row.is_opensource_api,
                stable_identity=sid,
                mapped_type=mapped_type,
                validation_result=vresult,
                validation_detail=vdetail or None,
                operational_state="duplicate" if is_duplicate else "inactive",
                inactive_reason="duplicate_identity" if is_duplicate else None,
            )
            session.add(inv)
            report.upserted += 1

        if vresult not in ("ok", "duplicate"):
            report.invalid.append(
                {
                    "workbook_id": row.workbook_id,
                    "name": row.name,
                    "platform": row.platform,
                    "result": vresult,
                    "detail": vdetail,
                }
            )

    report.platform_counts = platform_counts
    report.validation_counts = validation_counts
    session.flush()
    return report


# ── Activation ────────────────────────────────────────────────────


@dataclass
class ActivationReport:
    total: int = 0
    active: int = 0
    inactive: int = 0
    invalid: int = 0
    by_platform_state: dict[str, dict[str, int]] = field(default_factory=dict)
    inactive_reasons: dict[str, int] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "total": self.total,
            "active": self.active,
            "inactive": self.inactive,
            "invalid": self.invalid,
            "by_platform_state": self.by_platform_state,
            "inactive_reasons": self.inactive_reasons,
        }


def _activation_reason(
    inventory: SourceInventory,
    *,
    x_auth_available: bool,
    telegram_mtproto_available: bool,
) -> str | None:
    """Return an inactive_reason if the row must be inactive, else None."""
    if (
        inventory.operational_state == "inactive"
        and inventory.inactive_reason in PERMANENT_RUNTIME_INACTIVE_REASONS
    ):
        return inventory.inactive_reason
    if inventory.validation_result == "duplicate":
        return "duplicate_identity"
    if inventory.validation_result != "ok":
        return f"invalid:{inventory.validation_result}"
    mapped, reason = mapped_type_for(inventory.platform, inventory.workbook_type, inventory.public_url)
    if reason:
        return reason
    if mapped == "x_timeline" and not x_auth_available:
        return "x_auth_not_configured"
    if mapped == "telegram" and not telegram_mtproto_available:
        return "mtproto_not_configured"
    return None


def _unique_source_name(session: Session, base: str, workbook_id: int) -> str:
    """Disambiguate a source name (workbook may have duplicate display names)."""
    candidate = base or f"source-{workbook_id}"
    if not session.query(Source).filter_by(name=candidate).first():
        return candidate
    return f"{candidate} [#{workbook_id}]"


def activate_inventory_sources(
    session: Session,
    *,
    x_auth_available: bool | None = None,
    telegram_mtproto_available: bool | None = None,
) -> ActivationReport:
    """Activate accessible inventory sources into the ``sources`` registry.

    For each inventory row:
      * if inaccessible/invalid → operational_state=inactive with a reason
        (no sources row, or existing sources row disabled);
      * if accessible → upsert a ``sources`` row (enabled=True), link source_id.

    Activation is a DB-only step (no network). ``x_auth_available`` and
    ``telegram_mtproto_available`` default to environment-driven checks.
    """
    from newsroom.config import settings

    if x_auth_available is None:
        x_auth_available = bool(
            os.environ.get("TWITTER_AUTH_TOKEN", "").strip()
            and os.environ.get("TWITTER_CT0", "").strip()
        )
    if telegram_mtproto_available is None:
        telegram_mtproto_available = settings.telegram_ingestor_ready()

    report = ActivationReport()
    rows = session.query(SourceInventory).order_by(SourceInventory.workbook_id).all()
    report.total = len(rows)

    for inv in rows:
        reason = _activation_reason(
            inv,
            x_auth_available=x_auth_available,
            telegram_mtproto_available=telegram_mtproto_available,
        )
        plat = inv.platform
        ps = report.by_platform_state.setdefault(
            plat, {"active": 0, "inactive": 0, "invalid": 0, "duplicate": 0}
        )

        if reason is not None:
            inv.inactive_reason = reason
            if reason == "duplicate_identity":
                inv.operational_state = "duplicate"
                report.inactive += 1
                ps["duplicate"] += 1
                report.inactive_reasons[reason] = report.inactive_reasons.get(reason, 0) + 1
            elif reason.startswith("invalid:"):
                inv.operational_state = "invalid"
                report.invalid += 1
                ps["invalid"] += 1
            else:
                inv.operational_state = "inactive"
                report.inactive += 1
                ps["inactive"] += 1
                report.inactive_reasons[reason] = report.inactive_reasons.get(reason, 0) + 1
            # Disable any existing linked source without removing it (preserve history).
            if inv.source_id:
                src = session.get(Source, inv.source_id)
                if src and src.enabled:
                    src.enabled = False
                    src.inactive_reason = reason
            continue

        # Accessible — upsert sources row.
        src = (
            session.query(Source).filter_by(stable_identity=inv.stable_identity).first()
            if inv.stable_identity
            else None
        )
        if src is None and inv.source_id:
            src = session.get(Source, inv.source_id)
        if src is None:
            name = _unique_source_name(session, inv.name, inv.workbook_id)
            src = Source(
                name=name,
                type=inv.mapped_type,
                url=inv.public_url,
                language=(inv.language or "en")[:10],
                category=(inv.topic or "general")[:100],
                trust_class="reputable",
                enabled=True,
                config={},
                stable_identity=inv.stable_identity,
                workbook_id=inv.workbook_id,
                platform=inv.platform,
                inactive_reason=None,
                health_status="configured",
            )
            session.add(src)
            session.flush()
            inv.source_id = src.id
        else:
            src.type = inv.mapped_type
            src.url = inv.public_url
            src.language = (inv.language or "en")[:10]
            src.category = (inv.topic or src.category or "general")[:100]
            src.enabled = True
            src.stable_identity = inv.stable_identity
            src.workbook_id = inv.workbook_id
            src.platform = inv.platform
            src.inactive_reason = None
            if src.health_status not in ("healthy", "degraded"):
                src.health_status = "configured"
            inv.source_id = src.id

        inv.operational_state = "active"
        inv.inactive_reason = None
        report.active += 1
        ps["active"] += 1

    session.flush()
    return report


def reconciliation_summary(session: Session) -> dict[str, Any]:
    """Return a reconciliation summary of the current inventory state."""
    total = session.query(SourceInventory).count()
    by_platform: dict[str, dict[str, int]] = {}
    by_state: dict[str, int] = {}
    inactive_reasons: dict[str, int] = {}
    rows = session.query(SourceInventory).all()
    for inv in rows:
        ps = by_platform.setdefault(inv.platform, {"active": 0, "inactive": 0, "invalid": 0})
        ps[inv.operational_state] = ps.get(inv.operational_state, 0) + 1
        by_state[inv.operational_state] = by_state.get(inv.operational_state, 0) + 1
        if inv.inactive_reason:
            inactive_reasons[inv.inactive_reason] = inactive_reasons.get(inv.inactive_reason, 0) + 1
    accounted = sum(by_state.values())
    return {
        "total": total,
        "accounted": accounted,
        "reconciled": total == accounted,
        "by_platform": by_platform,
        "by_state": by_state,
        "inactive_reasons": inactive_reasons,
    }


__all__ = [
    "AUTHORITATIVE_SHEET",
    "IMPORT_DEST",
    "SOURCE_WORKBOOK_ENV",
    "ActivationReport",
    "ImportReport",
    "ParsedRow",
    "activate_inventory_sources",
    "copy_workbook_to_import_dir",
    "find_workbook",
    "import_workbook",
    "mapped_type_for",
    "reconciliation_summary",
    "stable_identity_for",
]
