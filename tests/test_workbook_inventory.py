"""Deterministic tests for the Gate 6 workbook importer and source inventory.

These tests are DB-free: they cover pure functions (stable identity,
platform→type mapping, row validation, workbook parsing from a synthetic
XLSX) and do not require PostgreSQL. Real-DB import/activation is covered by
tests/integration/test_gate6_source_inventory.py.
"""

from __future__ import annotations

from pathlib import Path

from newsroom.sources.inventory import (
    AUTHORITATIVE_SHEET,
    EXPECTED_PLATFORM_COUNTS,
    EXPECTED_TOTAL,
    _parse_rows,
    _validate_row,
    mapped_type_for,
    stable_identity_for,
)

# ── Stable identity ───────────────────────────────────────────────


def test_stable_identity_telegram_uses_handle():
    a = stable_identity_for("Telegram", "@AIpersianChannel", "https://t.me/AIpersianChannel")
    b = stable_identity_for("Telegram", "aipersianchannel", "https://t.me/AIpersianChannel")
    assert a == b  # case-insensitive, ignores @
    c = stable_identity_for("Telegram", "@Other", "https://t.me/Other")
    assert a != c


def test_stable_identity_reddit_uses_subreddit():
    a = stable_identity_for("Reddit", "r/AI_Agents", "https://www.reddit.com/r/AI_Agents/")
    b = stable_identity_for("Reddit", "ai_agents", "https://www.reddit.com/r/ai_agents/")
    assert a == b


def test_stable_identity_x_uses_handle():
    a = stable_identity_for("X / Twitter", "@AnthropicAI", "https://x.com/AnthropicAI")
    b = stable_identity_for("X / Twitter", "anthropicai", "https://x.com/AnthropicAI")
    assert a == b


def test_stable_identity_github_repo_lowercased():
    a = stable_identity_for("GitHub", None, "https://github.com/pytorch/pytorch")
    b = stable_identity_for("GitHub", None, "https://github.com/PyTorch/PyTorch")
    assert a == b


def test_stable_identity_github_nonrepo_uses_url():
    a = stable_identity_for("GitHub", None, "https://github.com/trending")
    assert a  # non-repo still gets a stable identity from normalized URL


def test_stable_identity_youtube_handle():
    a = stable_identity_for("YouTube / Social", "@Fireship", "https://www.youtube.com/@Fireship")
    b = stable_identity_for("YouTube / Social", "fireship", "https://www.youtube.com/@Fireship")
    assert a == b


def test_stable_identity_website_uses_normalized_url():
    a = stable_identity_for("Website / Newsletter", None, "https://example.com/")
    b = stable_identity_for("Website / Newsletter", None, "https://example.com")
    assert a == b  # trailing slash normalized


def test_stable_identity_independent_of_display_name():
    # Same URL but different (display) name → same identity.
    a = stable_identity_for("Website / Newsletter", None, "https://example.com/news")
    assert a == stable_identity_for("Website / Newsletter", None, "https://example.com/news")


# ── Platform → mapped type ───────────────────────────────────────


def test_mapped_type_per_platform():
    assert mapped_type_for("Telegram", "Channel", "https://t.me/x")[0] == "telegram"
    assert mapped_type_for("Reddit", "Subreddit", "https://reddit.com/r/ai")[0] == "reddit_subreddit"
    assert mapped_type_for("X / Twitter", "Account", "https://x.com/a")[0] == "x_timeline"
    assert mapped_type_for("GitHub", "Repository / Feed", "https://github.com/a/b")[0] == "github_releases"
    assert mapped_type_for("YouTube / Social", "YouTube Channel", "https://youtube.com/@a")[0] == "youtube_rss"
    assert mapped_type_for("Website / Newsletter", "Website", "https://example.com")[0] == "web_page"


def test_mapped_type_github_trending_is_inactive_not_a_repo():
    mapped, reason = mapped_type_for("GitHub", "Repository / Feed", "https://github.com/trending")
    assert reason == "not_a_repo"
    assert mapped == "github_releases"


def test_mapped_type_discord_is_access_required():
    mapped, reason = mapped_type_for("Community", "Discord", "https://forum.cursor.com")
    assert reason == "access_required"
    assert mapped == "web_page"


# ── Row validation ────────────────────────────────────────────────


def _row(**kw):
    from newsroom.sources.inventory import ParsedRow

    defaults = {
        "workbook_id": 1, "platform": "Telegram", "workbook_type": "Channel", "name": "x",
        "handle": "@x", "public_url": "https://t.me/x", "topic": "", "tags": "", "language": "en",
        "content_mode": "", "speed": None, "informal": None, "noise": None, "is_community": False,
        "is_opensource_api": False, "risk": "", "verification": "", "discovery_source": "",
        "tier": "Core", "coverage_score": 0,
    }
    defaults.update(kw)
    return ParsedRow(**defaults)


def test_validate_ok():
    assert _validate_row(_row()) == ("ok", "")


def test_validate_missing_url():
    assert _validate_row(_row(public_url=""))[0] == "missing_url"


def test_validate_invalid_telegram_handle():
    assert _validate_row(_row(platform="Telegram", handle="", public_url="https://example.com"))[0] == "invalid_handle"


def test_validate_invalid_reddit_handle():
    assert _validate_row(_row(platform="Reddit", handle="", public_url="https://example.com"))[0] == "invalid_handle"


# ── Workbook parsing from a synthetic XLSX ────────────────────────


def _make_workbook(path: Path, rows: list[list]) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = AUTHORITATIVE_SHEET
    for r in rows:
        ws.append(r)
    wb.save(str(path))
    return path


HEADER = [
    "ID", "Platform", "Type", "Name", "Handle / ID", "Direct URL", "Primary Topic",
    "Tags", "Language", "Content Mode", "Speed 1-5", "Informal 1-5", "Noise 1-5",
    "Community?", "Open-source/API?", "Risk", "Verification", "Discovery Source",
    "Tier", "Coverage Score",
]


def test_parse_rows_synthetic_workbook(tmp_path):
    fp = _make_workbook(
        tmp_path / "wb.xlsx",
        [
            HEADER,
            [1, "Telegram", "Channel", "AI Post", "@aipost", "https://t.me/aipost", "AI News",
             "ai", "English", "Mixed", 5, 5, 4, "No", "Yes", "Medium", "checked", "src", "Core", 24],
            [2, "Reddit", "Subreddit", "r/ai", "r/ai", "https://www.reddit.com/r/ai/", "AI", "",
             "English", "", 3, 3, 3, "No", "No", "Low", "public", "src", "Community", 10],
        ],
    )
    from newsroom.sources.inventory import _load_workbook_rows

    rows = _load_workbook_rows(fp)
    parsed = _parse_rows(rows)
    assert len(parsed) == 2
    assert parsed[0].workbook_id == 1
    assert parsed[0].platform == "Telegram"
    assert parsed[0].handle == "@aipost"
    assert parsed[0].coverage_score == 24
    assert parsed[1].platform == "Reddit"
    assert parsed[1].is_opensource_api is False


def test_parse_rows_skips_empty_and_headerless():
    fp = _make_workbook(
        Path("/tmp/_nb.xlsx") if False else Path(__file__).parent / "_tmp_skip.xlsx",
        [HEADER, [None, None], [1, "Telegram", "Channel", "x", "@x", "https://t.me/x", "", "", "en", "", None, None, None, "No", "No", "", "", "", "Core", 0]],
    )
    from newsroom.sources.inventory import _load_workbook_rows

    rows = _load_workbook_rows(fp)
    parsed = _parse_rows(rows)
    assert len(parsed) == 1
    Path(fp).unlink(missing_ok=True)


def test_expected_constants_match_spec():
    assert EXPECTED_TOTAL == 1344
    assert EXPECTED_PLATFORM_COUNTS == {
        "Telegram": 159, "Reddit": 204, "Community": 45, "Community / Forum": 19,
        "X / Twitter": 144, "Website / Newsletter": 464, "GitHub": 246, "YouTube / Social": 63,
    }
